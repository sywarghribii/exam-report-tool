"""
excel_export.py
===============

Excel output layer.  Consumes the typed models / flat rows from the rest of
the toolkit and produces a formatted ``.xlsx`` workbook:

* one sheet per report file, with the **Fails** column (all failing subtests)
  and the reporting columns (**Reporting Comment**, **Rerun**),
* an optional **Fail Analysis** sheet with the signature clusters, and
* an optional **Summary** sheet with the headline counts.

Formatting is deliberately light and dependency-free (openpyxl only): a bold
header row, frozen panes, wrapped text for the long columns and auto-sized
widths.  No formulas are written, so no recalculation step is required.
"""

from __future__ import annotations

import pathlib
from typing import Dict, Iterable, List, Optional, Union

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analyzer import cluster_by_signature, clusters_to_rows, failure_overview
from .comments import CommentStore, merge_comments
from .extractor import DEFAULT_METADATA_LABELS, report_to_rows
from .models import ExamReport

PathLike = Union[str, pathlib.Path]

# Column order for the per-report sheets.  Missing columns are skipped.
DEFAULT_COLUMN_ORDER = [
    "TestCaseID",
    "Name",
    "StartTime",
    "Duration",
    "InitialValuation",
    "FinalValuation",
    "Type",
    "GroupPath",
    "FailCount",
    "Fails",
    "Variante",
    "Platform",
    "Projekt",
    "Link",
    "Reporting Comment",
    "Rerun",
]

# Columns that hold long text and should wrap.
_WRAP_COLUMNS = {"Fails", "Reporting Comment", "Name", "Signature", "Example", "TestCaseIDs"}

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_BODY_FONT = Font(name="Arial")
_MAX_WIDTH = 60  # cap so a giant Fails cell doesn't blow the layout out


def _sanitize_sheet_name(name: str, used: set) -> str:
    """Excel sheet names: <=31 chars, no ``[]:*?/\\`` and must be unique."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    name = name[:31] or "Sheet"
    base, i = name, 1
    while name in used:
        suffix = f"_{i}"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def _as_text(value) -> str:
    """Safely stringify a cell value (NaN/None -> empty string).

    pandas' newer string dtype can carry NaN through ``astype(str)``, so we
    normalise explicitly instead of trusting the dtype conversion.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value)


def _write_dataframe(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    """Write *df* to a sheet and apply the shared formatting."""
    # Fill missing values so optional columns don't leave stray NaN cells.
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == object or str(df[col].dtype) == "str":
            df[col] = df[col].map(_as_text)
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    # Header styling + freeze panes.
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    # Body font + wrapping + column widths.
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        wrap = col_name in _WRAP_COLUMNS
        # width from the longest line in any cell (respecting embedded newlines)
        longest = len(str(col_name))
        for value in df[col_name]:
            for line in _as_text(value).split("\n"):
                longest = max(longest, len(line))
        ws.column_dimensions[letter].width = min(longest + 2, _MAX_WIDTH)
        for row_idx in range(2, len(df) + 2):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = _BODY_FONT
            if wrap:
                c.alignment = Alignment(wrap_text=True, vertical="top")


def export_reports_to_excel(
    reports: Dict[str, ExamReport],
    excel_path: PathLike,
    comment_store: Optional[CommentStore] = None,
    metadata_labels: Iterable[str] = DEFAULT_METADATA_LABELS,
    include_analysis: bool = True,
    include_summary: bool = True,
    column_order: Optional[List[str]] = None,
) -> pathlib.Path:
    """Export one or more parsed reports to a formatted workbook.

    *reports* maps a label (usually the file path) to an :class:`ExamReport`.
    If *comment_store* is given, its comments are merged into the reporting
    columns.  Set *include_analysis* / *include_summary* to add the extra
    cross-report sheets.
    """
    excel_path = pathlib.Path(excel_path)
    column_order = column_order or DEFAULT_COLUMN_ORDER
    used_names: set = set()

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # ---- optional summary sheet (written first so it's the landing tab)
        if include_summary:
            overview = failure_overview(reports.values())
            summary_df = pd.DataFrame(
                [{"Metric": k, "Value": v} for k, v in overview.items()]
            )
            _write_dataframe(writer, summary_df, _sanitize_sheet_name("Summary", used_names))

        # ---- one sheet per report
        for label, report in reports.items():
            rows = report_to_rows(report, metadata_labels=metadata_labels)
            if comment_store is not None:
                merge_comments(rows, comment_store)
            df = pd.DataFrame(rows)
            ordered = [c for c in column_order if c in df.columns]
            extras = [c for c in df.columns if c not in ordered]
            df = df[ordered + extras]
            sheet = _sanitize_sheet_name(pathlib.Path(label).stem, used_names)
            _write_dataframe(writer, df, sheet)

        # ---- optional fail-analysis sheet
        if include_analysis:
            clusters = cluster_by_signature(reports.values())
            analysis_df = pd.DataFrame(clusters_to_rows(clusters))
            if not analysis_df.empty:
                _write_dataframe(
                    writer, analysis_df, _sanitize_sheet_name("Fail Analysis", used_names)
                )

    return excel_path


def export_files_to_excel(
    xml_paths: Iterable[PathLike],
    excel_path: PathLike,
    comment_store_path: Optional[PathLike] = None,
    **kwargs,
) -> pathlib.Path:
    """Convenience one-liner: parse files and export in a single call.

    >>> export_files_to_excel(["a.xml", "b.xml"], "out.xlsx")
    """
    from .extractor import parse_reports

    reports = parse_reports(xml_paths)
    store = CommentStore.load(comment_store_path) if comment_store_path else None
    return export_reports_to_excel(reports, excel_path, comment_store=store, **kwargs)
